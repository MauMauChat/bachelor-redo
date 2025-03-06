#!/usr/bin/env python3
import logging
import datetime
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

class ExcelWriter:
    def __init__(self, template_path, output_path):
        self.template_path = template_path
        self.output_path = output_path
        logging.info(f"ExcelWriter initialisiert mit Template: {template_path}, Output: {output_path}")

    def _ermittle_semester(self):
        heute = datetime.date.today()
        year = heute.year
        month = heute.month
        day = heute.day

        if (month > 2 and month < 11) or (month == 3 and day >= 1) or (month == 10 and day <= 30):
            return year, "Sommer"
        else:
            return year, "Winter"

    def write_rows(self, rows):
        """
        rows ist eine Liste aus Dictionaries mit Keys ["i", "s", "c"], also:
          [
            {"i": "1", "s": "Satz A", "c": "Positiv"},
            {"i": "2", "s": "Satz B", "c": "Bezug auf LV"},
            ...
          ]
        """
        try:
            # 1) Excel-Vorlage laden
            wb = openpyxl.load_workbook(self.template_path)
            sheet = wb.active

            # 2) Jahr & Semester bestimmen
            jahr, semester = self._ermittle_semester()
            logging.info(f"Ermitteltes Semester: {semester} {jahr}")

            sheet["A1"] = "Semester"
            sheet["B1"] = semester
            sheet["A2"] = "Jahr"
            sheet["B2"] = jahr

            # 3) Mapping: Kategorien -> Liste von (ID, Satz)
            category_list = [
                "Bezug auf LV",
                "Nicht verständlich",
                "Nicht eindeutig zuordenbar",
                "Negativ",
                "Positiv",
                "Neutral, kein Kommentar, äquivalente Symbole, bedeutungslos",
                "Anregungen, Wünsche"
            ]
            # Jede Kategorie bekommt ein leeres Array, in das wir Tuples (ID, Satz) speichern
            results_by_category = {cat: [] for cat in category_list}

            # Gleichzeitig wollen wir alles nochmal in die Excel-Tabelle schreiben (Spalten A, B, C)
            start_header = 16
            sheet.cell(row=start_header, column=1, value="ID")
            sheet.cell(row=start_header, column=2, value="Kommentar")
            sheet.cell(row=start_header, column=3, value="Kategorie")

            data_row = start_header + 1
            for row_data in rows:
                i_val = row_data["i"]
                s_val = row_data["s"]
                c_val = row_data["c"]

                # In die Excel-Tabelle schreiben
                sheet.cell(row=data_row, column=1, value=i_val)   # ID
                sheet.cell(row=data_row, column=2, value=s_val)     # Kommentar
                sheet.cell(row=data_row, column=3, value=c_val)     # Kategorie
                data_row += 1

                # Mapping befüllen
                if c_val in results_by_category:
                    results_by_category[c_val].append((i_val, s_val))

            # 4) Kategorie-Zählungen ermitteln
            cat_counts = {cat: len(results_by_category[cat]) for cat in category_list}
            total = sum(cat_counts.values())

            # 5) Tortendiagramm erstellen
            plt.figure()
            plt.pie(
                cat_counts.values(),
                labels=cat_counts.keys(),
                autopct='%1.1f%%',
                startangle=140
            )
            plt.title(f"Auswertung {semester}semester {jahr}")
            plt.axis('equal')  # Kreisdiagramm

            # Diagramm speichern
            diagramm_name = f"auswertung_{semester}_{jahr}.png"
            plt.savefig(diagramm_name, bbox_inches="tight")
            plt.close()
            logging.info(f"Tortendiagramm gespeichert: {diagramm_name}")

            # 6) Diagramm in Excel einfügen
            img = OpenpyxlImage(diagramm_name)
            # Beispiel: Bild in Zelle H2 platzieren
            sheet.add_image(img, "H2")

            # 7) Summen ab Zeile 19 in Spalte E/F einfügen
            start_summary = 19
            for i, cat in enumerate(category_list):
                row_index = start_summary + i
                sheet.cell(row=row_index, column=5, value=cat)            # Kategorie
                sheet.cell(row=row_index, column=6, value=cat_counts[cat])  # Anzahl Sätze

            # Gesamtergebnis in Zeile 26
            total_row = start_summary + len(category_list)
            sheet.cell(row=total_row, column=5, value="Gesamtergebnis")
            sheet.cell(row=total_row, column=6, value=total)

            # 8) Excel speichern
            wb.save(self.output_path)
            logging.info(f"Excel-Auswertung erstellt: {self.output_path}")

        except Exception as e:
            logging.error(f"Fehler beim Schreiben in Excel-Template: {e}")
