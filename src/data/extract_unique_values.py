#!/usr/bin/env python3
import os

import pandas as pd
import openpyxl


def read_csv_file(file_path, delimiter=';', encoding='utf-8'):
    """
    Liest eine CSV-Datei ein – fehlerhafte Zeilen werden übersprungen.
    """
    try:
        df = pd.read_csv(file_path,
                         delimiter=delimiter,
                         encoding=encoding,
                         on_bad_lines='skip',
                         low_memory=False)
        print(f"{file_path} erfolgreich gelesen: {len(df)} Zeilen")
    except Exception as e:
        print(f"Fehler beim Lesen von {file_path}: {e}")
        df = pd.DataFrame()
    return df


def compute_statistics(df_frageboegen, df_fragen, df_freitext):
    """
    Ermittelt aus den drei CSVs die Kennzahlen, die in das Excel‑Template eingetragen werden sollen.

    Dabei werden für manche Kennzahlen (z. B. Anzahl der LVs) eindeutige Werte (unique) extrahiert.

    Passen Sie die Logik (GroupBy, Summen, Zählungen etc.) hier exakt an, damit die aus den CSV‑Dateien
    extrahierten Werte Ihren Anforderungen entsprechen!
    """
    data = {}

    # --- Beispiel: Anzahl der LVs (Template: "Anzahl der LVs:")
    # Hier wird aus der Datei "Alle Fragebögen.csv" die eindeutige Anzahl der LVs ermittelt.
    df_frageboegen.columns = df_frageboegen.columns.str.strip("'")

    print(df_frageboegen.columns.tolist())

    if 'RLV_KEY' in df_frageboegen.columns:
        data["Anzahl der LVs:"] = df_frageboegen['RLV_KEY'].nunique()
    elif 'RLVKEY' in df_frageboegen.columns:
        data["Anzahl der LVs:"] = df_frageboegen['RLVKEY'].nunique()
    else:
        data["Anzahl der LVs:"] = 0

    # --- Beispiel: Ausgefüllte Fragebögen (z. B. eindeutige FRAGE_KEY in "Alle Fragebögen.csv")
    if 'FRAGE_KEY' in df_frageboegen.columns:
        data["Ausgefüllte Fragebögen:"] = df_frageboegen['FRAGE_KEY'].nunique()
    else:
        data["Ausgefüllte Fragebögen:"] = 0

    # --- Beispiel: Abgelehnte Fragebögen (Summe der eindeutigen ANZ_ABGELEHNT pro LV)
    if 'ANZ_ABGELEHNT' in df_frageboegen.columns and 'RLV_KEY' in df_frageboegen.columns:
        # Gruppieren nach LV und Summe der abgelehnten Fragebögen bilden
        grp = df_frageboegen.groupby('RLV_KEY')['ANZ_ABGELEHNT'].apply(
            lambda x: pd.to_numeric(x, errors='coerce').sum())
        data["Abgelehnte Fragebögen:"] = grp.sum()
    else:
        data["Abgelehnte Fragebögen:"] = 0

    # --- Beispiel: LVs ohne Anmeldungen (in "Alle Fragebögen.csv": ANZ_RUECKLAUF == 0)
    if 'ANZ_RUECKLAUF' in df_frageboegen.columns and 'RLV_KEY' in df_frageboegen.columns:
        grp = df_frageboegen.groupby('RLV_KEY')['ANZ_RUECKLAUF'].apply(
            lambda x: pd.to_numeric(x, errors='coerce').sum())
        data["LVs ohne Anmeldungen:"] = (grp == 0).sum()
    else:
        data["LVs ohne Anmeldungen:"] = 0

    # --- Beispiel: LVs mit Anmeldungen aber ohne Aufnahmen
    # (Hier wird angenommen, dass "ANZ_ANTWORT_FRAGE_GESAMT" in "Alle Fragebögen.csv" herangezogen wird)
    if all(col in df_frageboegen.columns for col in ['ANZ_RUECKLAUF', 'ANZ_ANTWORT_FRAGE_GESAMT', 'RLV_KEY']):
        grp = df_frageboegen.groupby('RLV_KEY').agg({
            'ANZ_RUECKLAUF': lambda x: pd.to_numeric(x, errors='coerce').sum(),
            'ANZ_ANTWORT_FRAGE_GESAMT': lambda x: pd.to_numeric(x, errors='coerce').sum()
        })
        data["LVs mit Anmeldungen aber ohne Aufnahmen:"] = (
                    (grp['ANZ_RUECKLAUF'] > 0) & (grp['ANZ_ANTWORT_FRAGE_GESAMT'] == 0)).sum()
    else:
        data["LVs mit Anmeldungen aber ohne Aufnahmen:"] = 0

    # --- Beispiel: LVs zu denen Feedback gegeben werden konnte (ANZ_FEEDBACK > 0)
    if 'ANZ_FEEDBACK' in df_frageboegen.columns and 'RLV_KEY' in df_frageboegen.columns:
        grp = df_frageboegen.groupby('RLV_KEY')['ANZ_FEEDBACK'].apply(
            lambda x: pd.to_numeric(x, errors='coerce').sum())
        data["LVs zu denen Feedback gegeben werden konnte:"] = (grp > 0).sum()
    else:
        data["LVs zu denen Feedback gegeben werden konnte:"] = 0

    # --- Beispiel: Bewertung – hier aus der Datei mit Freitextantworten
    # Angenommen, die Spalte "NOTE" enthält Bewertungen (aus "FB Freitextantworten.csv")
    if 'NOTE' in df_freitext.columns and 'RLVKEY' in df_freitext.columns:
        # Eindeutige Zuordnung von LV zu Bewertung (hier: erster gültiger Eintrag pro LV)
        df_tmp = df_freitext[['RLVKEY', 'NOTE']].dropna()
        df_tmp['NOTE'] = pd.to_numeric(df_tmp['NOTE'], errors='coerce')
        # Beispiel: LVs mit Gesamtbewertung gleich 1
        data["LVs mit Gesamtbewertung gleich 1:"] = (df_tmp['NOTE'] == 1).sum()
    else:
        data["LVs mit Gesamtbewertung gleich 1:"] = 0

    # --- Weitere Metriken aus dem Template
    # Falls für bestimmte Metriken keine direkte Berechnung definiert ist, wird hier 0 eingetragen.
    template_keys = [
        "Berechtigte Studierende:",
        "Feedbackberechtigungen:",
        "Prozentsatz der Beteiligung:",
        "LVs ohne Bewertung:",
        "LVs ohne Bewertung wegen Nichtberechtigung:",
        "LVs mit einer Bewertung:",
        "LVs mit zwei Bewertungen:",
        "LVs mit drei Bewertungen:",
        "LVs mit mehr als 3 Bewertungen:",
        "LVs mit Feedback:",
        "LVs mit 1-3 Bewertungen:",
        "LVs, deren Veröffentlichung nicht zugestimmt wurde:",
        "Anzahl der LV-Leiter/innen:",
        "LV-Leiter/innen, die Feedback erhalten haben:",
        "LV-Leiter/innen, die individuelle Komponenten erstellt haben:",
        "LV-Leiter/innen, die eine Stellungnahme abgegeben haben:",
        "Stellungnahmen zu veröffentlichten Ergebnissen:",
        "Stellungnahmen zu nicht veröffentlichten Ergebnissen:",
        "LV-Leiter/innen, die der Veröffentlichung nicht zugestimmt haben:",
        "LVs mit Gesamtbewertung besser oder gleich 1,2:",
        "LVs mit Gesamtbewertung besser oder gleich 1,5:",
        "LVs mit Gesamtbewertung besser oder gleich 2:",
        "LVs mit Gesamtbewertung besser oder gleich 2,5:",
        "LVs mit Gesamtbewertung schlechter oder gleich 2,5:",
        "LVs mit Gesamtbewertung schlechter oder gleich 3:",
        "LVs mit Gesamtbewertung schlechter oder gleich 3,5:",
        "LVs mit Gesamtbewertung schlechter oder gleich 4:"
    ]
    for key in template_keys:
        if key not in data:
            data[key] = 0

    return data


def fill_template(data, template_file, output_file):
    """
    Lädt die Excel‑Vorlage, sucht in Spalte A nach den Template‑Kennzahlen und trägt in Spalte B
    den zugehörigen (berechneten) Wert ein – falls kein Wert vorliegt, wird 0 eingetragen.
    Die Formatierung der Vorlage bleibt erhalten.
    """
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active

    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val is not None:
            metric = cell_val.strip()
            if metric in data:
                ws.cell(row=row, column=2, value=data[metric])
            else:
                ws.cell(row=row, column=2, value=0)


    if os.path.exists(output_file):
        os.remove(output_file)
    wb.save(output_file)

    print(f"Die Statistik wurde in '{output_file}' gespeichert.")


def main():
    # --- CSV-Dateipfade (bitte anpassen) ---
    path_frageboegen = "/home/lucy/PycharmProjects/bachelorarbeit_redo/project/src/data/Alle Fragebögen.csv"
    path_fragen = "/home/lucy/PycharmProjects/bachelorarbeit_redo/project/src/data/Alle Fragen.csv"
    path_freitext = "/home/lucy/PycharmProjects/bachelorarbeit_redo/project/src/data/FB Freitextantworten.csv"

    df_frageboegen = read_csv_file(path_frageboegen)
    df_fragen = read_csv_file(path_fragen)
    df_freitext = read_csv_file(path_freitext)

    # --- Statistik berechnen (die Funktion passt die Daten aus den CSVs an die Template‑Merkmale an) ---
    data = compute_statistics(df_frageboegen, df_fragen, df_freitext)

    # --- Excel‑Template (mit Formatierung) und Ausgabe-Datei ---
    template_file = "/home/lucy/PycharmProjects/bachelorarbeit_redo/project/templates/Aggregierte-Statistik-Template.xlsx"
    output_file = "Feedback_Statistik_Ergebnis.xlsx"

    fill_template(data, template_file, output_file)


if __name__ == "__main__":
    main()
